r"""
=============================================================================
Script Name: Options_Journal_Generator_V5.py
Purpose: Generates the v5 AUTOMATED Excel (.xlsx) journal.
         
         V5 UPGRADES (The "Macro vs Micro" Volatility Correction):
         - Explicitly separated "Macro VIX" and "Chain ATM IV" to respect 
           Term Structure and Volatility Skew per CIO Handover Part 6.
         - Added "Exit Chain ATM IV (%)" to track Micro Vega Crush.
         - Re-aligned all formulas and conditional formatting to accommodate
           the new Column Q, pushing Closing Price to Column R.
         - Maintained all #VALUE! error shielding and Gray Formula Backgrounds.

Author: Chief Investment Officer AI Advisor
Date: April 2026
=============================================================================
"""

import os
import sys
import subprocess

# Auto-Install Missing Libraries
def install_packages():
    required_packages = ['pandas', 'openpyxl']
    for package in required_packages:
        try:
            __import__(package)
        except ImportError:
            print(f"Missing library '{package}' detected. Installing now...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])

install_packages()

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# Define Paths
target_directory = r"C:\Users\donca\Desktop\Desktop HP Envy x360 al 22Abr24\Docs Manuel\IBKR_Options"
file_name = "XSP_XND_Options_Journal_v5.xlsx"
full_file_path = os.path.join(target_directory, file_name)

if not os.path.exists(target_directory):
    os.makedirs(target_directory)

# Define Columns (Strict Macro vs Micro Separation)
columns =[
    "Tranche ID",                  # A
    "Open Date",                   # B
    "Ticker",                      # C 
    "Macro VIX at Entry",          # D (Global Sentiment)
    "Chain ATM IV at Entry (%)",   # E (Micro Pricing Physics)
    "DTE at Entry",                # F 
    "Short Strike",                # G
    "Long Strike",                 # H
    "Quantity",                    # I
    "Premium Collected (USD)",     # J
    "Collateral Locked (USD)",     # K (Formula)
    "Total Net Credit (USD)",      # L (Formula)
    "Target 50% Exit Price (USD)", # M (Formula)
    "Close Date",                  # N 
    "Days Remaining",              # O (Formula Countdown)
    "Exit Macro VIX",              # P 
    "Exit Chain ATM IV (%)",       # Q (NEW - Tracks exact Micro Vega Crush)
    "Closing Price (USD)",         # R (Shifted)
    "Days in Trade",               # S (Formula)
    "Total P&L (USD)",             # T (Formula)
    "Return on Capital (ROC) %",   # U (Formula)
    "Annualized ROC %",            # V (Formula)
    "Notes / Adjustments"          # W
]

try:
    df = pd.DataFrame(columns=columns)
    df.to_excel(full_file_path, index=False, engine='openpyxl')
    
    wb = load_workbook(full_file_path)
    ws = wb.active
    ws.title = "Options Ledger v5"
    
    # Base Formatting
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Gray fill for automated formula columns
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    formula_cols =['K', 'L', 'M', 'O', 'S', 'T', 'U', 'V']
    
    for col_num in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        ws.column_dimensions[get_column_letter(col_num)].width = 16

    ws.column_dimensions['A'].width = 12  
    ws.column_dimensions['W'].width = 30  

    # Inject Formulas & Gray Backgrounds
    for i in range(2, 102):
        # K: Collateral Locked
        ws[f'K{i}'] = f'=IFERROR(IF(OR(ISBLANK(G{i}),ISBLANK(H{i}),ISBLANK(I{i})),"",(G{i}-H{i})*100*I{i}), "")'
        ws[f'K{i}'].number_format = '#,##0.00'

        # L: Total Net Credit
        ws[f'L{i}'] = f'=IFERROR(IF(OR(ISBLANK(J{i}),ISBLANK(I{i})),"",J{i}*100*I{i}), "")'
        ws[f'L{i}'].number_format = '#,##0.00'

        # M: Target 50% Exit Price
        ws[f'M{i}'] = f'=IFERROR(IF(ISBLANK(J{i}),"",J{i}/2), "")'
        ws[f'M{i}'].number_format = '#,##0.00'

        # O: Days Remaining
        ws[f'O{i}'] = f'=IFERROR(IF(ISBLANK(B{i}), "", IF(ISBLANK(N{i}), F{i} - (TODAY() - B{i}), "Closed")), "")'
        ws[f'O{i}'].alignment = center_alignment

        # S: Days in Trade
        ws[f'S{i}'] = f'=IFERROR(IF(OR(ISBLANK(N{i}),ISBLANK(B{i})),"",N{i}-B{i}), "")'
        ws[f'S{i}'].number_format = '0'
        ws[f'S{i}'].alignment = center_alignment

        # T: Total P&L
        ws[f'T{i}'] = f'=IFERROR(IF(OR(ISBLANK(R{i}),ISBLANK(J{i}),ISBLANK(I{i})),"",(J{i}-R{i})*100*I{i}), "")'
        ws[f'T{i}'].number_format = '#,##0.00'

        # U: ROC %
        ws[f'U{i}'] = f'=IFERROR(IF(OR(ISBLANK(T{i}),ISBLANK(K{i}),K{i}=0),"",T{i}/K{i}), "")'
        ws[f'U{i}'].number_format = '0.00%'

        # V: Annualized ROC %
        ws[f'V{i}'] = f'=IFERROR(IF(OR(ISBLANK(U{i}),ISBLANK(S{i}),S{i}=0),"",U{i}*(365/S{i})), "")'
        ws[f'V{i}'].number_format = '0.00%'

        # Apply Gray Background to formula columns
        for col in formula_cols:
            ws[f'{col}{i}'].fill = gray_fill

        # Formats for inputs
        ws[f'J{i}'].number_format = '#,##0.00' # Premium Collected
        ws[f'R{i}'].number_format = '#,##0.00' # Closing Price
        ws[f'E{i}'].number_format = '0.00'     # Chain IV Entry
        ws[f'D{i}'].number_format = '0.00'     # Macro VIX Entry
        ws[f'P{i}'].number_format = '0.00'     # Macro VIX Exit
        ws[f'Q{i}'].number_format = '0.00'     # Chain IV Exit

    # ---------------------------------------------------------
    # Step 6: Conditional Formatting
    # ---------------------------------------------------------
    # 1. DAYS REMAINING (COLUMN O) - Time Stop Alarms
    safe_green_fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    safe_green_font = Font(bold=True, color="0F5132")
    ws.conditional_formatting.add('O2:O101',
        FormulaRule(formula=['AND(ISNUMBER(O2), O2>=(F2/2))'], stopIfTrue=True, fill=safe_green_fill, font=safe_green_font)
    )

    danger_red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    danger_red_font = Font(bold=True, color="842029")
    ws.conditional_formatting.add('O2:O101',
        FormulaRule(formula=['AND(ISNUMBER(O2), O2<(F2/2))'], stopIfTrue=True, fill=danger_red_fill, font=danger_red_font)
    )

    # 2. DAYS IN TRADE (COLUMN S) - Capital Velocity (Fast Exits)
    mint_green_fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    ws.conditional_formatting.add('S2:S101',
        CellIsRule(operator='lessThanOrEqual', formula=['14'], stopIfTrue=True, fill=mint_green_fill)
    )

    # 3. ANNUALIZED ROC (COLUMN V) - Astronomical Returns
    gold_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    bold_gold_font = Font(bold=True, color="856404")
    ws.conditional_formatting.add('V2:V101',
        CellIsRule(operator='greaterThan', formula=['1.0'], stopIfTrue=True, fill=gold_fill, font=bold_gold_font)
    )

    ws.freeze_panes = "A2"
    wb.save(full_file_path)
    
    print("\n========================================================")
    print(f"SUCCESS! Options Journal v5 (Macro/Micro Corrected) is ready.")
    print(f"Location: {full_file_path}")
    print("========================================================")

except PermissionError:
    print("\nERROR: Permission denied. Please close the Excel file if it is currently open.")
except Exception as e:
    print(f"\nAn unexpected error occurred: {e}")