r"""
=============================================================================
Script Name: Options_Journal_Generator_V3.py
Purpose: Generates the v3 AUTOMATED Excel (.xlsx) journal for tracking the 
         Family Office XSP and XND Put Credit Spread Ladder.
         
         V3 UPGRADES:
         - New Schema integration to support "Days Remaining" countdowns.
         - Added "Inst IV at Entry" and "Exit VIX/IV" to track Volatility Crush.
         - Automated Conditional Formatting for "Capital Velocity" (fast exits).
         - Exact column shifting to preserve the CIO's custom formulas.

Author: Chief Investment Officer AI Advisor
Date: April 2026
=============================================================================
"""

import os
import sys
import subprocess

# ---------------------------------------------------------
# Step 1: Auto-Install Missing Libraries
# ---------------------------------------------------------
def install_packages():
    """Checks for required libraries and installs them via pip if missing."""
    required_packages = ['pandas', 'openpyxl']
    for package in required_packages:
        try:
            __import__(package)
        except ImportError:
            print(f"Missing library '{package}' detected. Installing now...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])
            print(f"'{package}' successfully installed.")

install_packages()

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ---------------------------------------------------------
# Step 2: Define the Exact Directory and File Path
# ---------------------------------------------------------
target_directory = r"C:\Users\donca\Desktop\Desktop HP Envy x360 al 22Abr24\Docs Manuel\IBKR_Options"
# Renamed to v3 to prevent overwriting your legacy ledger just in case
file_name = "XSP_XND_Options_Journal_v3.xlsx"
full_file_path = os.path.join(target_directory, file_name)

if not os.path.exists(target_directory):
    os.makedirs(target_directory)
    print(f"Created new directory: {target_directory}")

# ---------------------------------------------------------
# Step 3: Define Columns and Create Base File
# ---------------------------------------------------------
# Schema shifted to accommodate user's exact formula requests + IV metrics
columns =[
    "Tranche ID",                  # A
    "Open Date",                   # B
    "Ticker",                      # C (XSP / XND)
    "Market VIX at Entry",         # D
    "Inst IV at Entry (%)",        # E (NEW: Instrument Specific IV)
    "DTE at Entry",                # F
    "Short Strike",                # G
    "Long Strike",                 # H
    "Quantity",                    # I
    "Premium Collected (USD)",     # J
    "Collateral Locked (USD)",     # K (Formula)
    "Total Net Credit (USD)",      # L (Formula)
    "Target 50% Exit Price (USD)", # M (Formula)
    "Close Date",                  # N
    "Days Remaining",              # O (Formula countdown)
    "Exit VIX/IV",                 # P (NEW: To measure Vega Crush)
    "Closing Price (USD)",         # Q
    "Days in Trade",               # R (Formula)
    "Total P&L (USD)",             # S (Formula)
    "Return on Capital (ROC) %",   # T (Formula)
    "Annualized ROC %",            # U (Formula)
    "Notes / Adjustments"          # V
]

try:
    # Generate the base file
    df = pd.DataFrame(columns=columns)
    df.to_excel(full_file_path, index=False, engine='openpyxl')
    
    # ---------------------------------------------------------
    # Step 4: Open the File with OpenPyXL to Inject Formulas
    # ---------------------------------------------------------
    wb = load_workbook(full_file_path)
    ws = wb.active
    ws.title = "Options Ledger v3"
    
    # Format Headers (Yellow Background, Bold, Centered)
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col_num in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        # Adjust column widths for readability
        ws.column_dimensions[get_column_letter(col_num)].width = 16

    # Specific width adjustments
    ws.column_dimensions['A'].width = 12  # Tranche ID
    ws.column_dimensions['V'].width = 30  # Notes

    # ---------------------------------------------------------
    # Step 5: Inject Dynamic Excel Formulas (Rows 2 to 100)
    # ---------------------------------------------------------
    for i in range(2, 102):
        # K: Collateral Locked = (Short Strike - Long Strike) * 100 * Quantity
        ws[f'K{i}'] = f'=IF(OR(ISBLANK(G{i}),ISBLANK(H{i}),ISBLANK(I{i})),"",(G{i}-H{i})*100*I{i})'
        ws[f'K{i}'].number_format = '#,##0.00'

        # L: Total Net Credit = Premium Collected * 100 * Quantity
        ws[f'L{i}'] = f'=IF(OR(ISBLANK(J{i}),ISBLANK(I{i})),"",J{i}*100*I{i})'
        ws[f'L{i}'].number_format = '#,##0.00'

        # M: Target 50% Exit Price = Premium Collected / 2
        ws[f'M{i}'] = f'=IF(ISBLANK(J{i}),"",J{i}/2)'
        ws[f'M{i}'].number_format = '#,##0.00'

        # O: Days Remaining = Dynamic Countdown utilizing TODAY()
        ws[f'O{i}'] = f'=IF(ISBLANK(B{i}), "", IF(ISBLANK(N{i}), F{i} - (TODAY() - B{i}), "Closed"))'
        ws[f'O{i}'].alignment = center_alignment

        # R: Days in Trade = Close Date - Open Date
        ws[f'R{i}'] = f'=IF(OR(ISBLANK(N{i}),ISBLANK(B{i})),"",N{i}-B{i})'
        ws[f'R{i}'].number_format = '0'
        ws[f'R{i}'].alignment = center_alignment

        # S: Total P&L = (Premium Collected - Closing Price) * 100 * Quantity
        ws[f'S{i}'] = f'=IF(OR(ISBLANK(Q{i}),ISBLANK(J{i}),ISBLANK(I{i})),"",(J{i}-Q{i})*100*I{i})'
        ws[f'S{i}'].number_format = '#,##0.00'

        # T: ROC % = Total P&L / Collateral Locked
        ws[f'T{i}'] = f'=IF(OR(ISBLANK(S{i}),ISBLANK(K{i}),K{i}=0),"",S{i}/K{i})'
        ws[f'T{i}'].number_format = '0.00%'

        # U: Annualized ROC % = ROC % * (365 / Days in Trade)
        ws[f'U{i}'] = f'=IF(OR(ISBLANK(T{i}),ISBLANK(R{i}),R{i}=0),"",T{i}*(365/R{i}))'
        ws[f'U{i}'].number_format = '0.00%'

        # Format input currency/decimal columns
        ws[f'J{i}'].number_format = '#,##0.00' # Premium
        ws[f'Q{i}'].number_format = '#,##0.00' # Closing Price
        ws[f'E{i}'].number_format = '0.00'     # Inst IV
        ws[f'P{i}'].number_format = '0.00'     # Exit VIX

    # ---------------------------------------------------------
    # Step 6: Inject Conditional Formatting (Capital Velocity)
    # ---------------------------------------------------------
    # Rule 1: Fast Exits (<= 14 days) -> Mint Green background
    mint_green_fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    ws.conditional_formatting.add('R2:R101',
        CellIsRule(operator='lessThanOrEqual', formula=['14'], stopIfTrue=True, fill=mint_green_fill)
    )

    # Rule 2: Astronomical Yields (> 100% Ann. ROC) -> Gold background, Bold
    gold_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    bold_font = Font(bold=True, color="856404")
    ws.conditional_formatting.add('U2:U101',
        CellIsRule(operator='greaterThan', formula=['1.0'], stopIfTrue=True, fill=gold_fill, font=bold_font)
    )

    # Freeze the top header row so it stays visible when scrolling down
    ws.freeze_panes = "A2"
    
    # Save the updated workbook
    wb.save(full_file_path)
    
    print("\n========================================================")
    print(f"SUCCESS! The V3 Options Journal & Analytics Engine is ready.")
    print(f"Location: {full_file_path}")
    print("========================================================")

except PermissionError:
    print("\nERROR: Permission denied. Please close the Excel file if it is currently open and run the script again.")
except Exception as e:
    print(f"\nAn unexpected error occurred: {e}")