"""
=============================================================================
Script Name: Options_Journal_Generator_V2.py
Purpose: Generates a FULLY AUTOMATED Excel (.xlsx) journal for tracking the 
         Family Office XSP Put Credit Spread Options Ladder.
         This version injects smart Excel formulas (Target Exits, P&L, 
         and Annualized ROC) directly into the cells.
Author: Chief Investment Officer AI Advisor
Date: March 2026

How to Run:
1. Ensure Python is installed on your Windows PC.
2. Save this script as 'Options_Journal_Generator_V2.py' in your designated folder:
   C:\\Users\\donca\\Desktop\\Desktop HP Envy x360 al 22Abr24\\Docs Manuel\\IBKR_Options
3. Open your Command Prompt (cmd) or run it directly from your Python IDE.
4. The script will automatically check for the required libraries.
5. The script will output the Excel template to your exact folder path, pre-loaded 
   with 100 rows of automated formulas and professional formatting.
=============================================================================
"""

import os
import sys
import subprocess

# ---------------------------------------------------------
# Step 1: Auto-Install Missing Libraries
# ---------------------------------------------------------
def install_packages():
    """Checks for required libraries and installs them via pip if missing."""
    required_packages =['pandas', 'openpyxl']
    for package in required_packages:
        try:
            __import__(package)
        except ImportError:
            print(f"Missing library '{package}' detected. Installing now...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])
            print(f"'{package}' successfully installed.")

install_packages()

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------
# Step 2: Define the Exact Directory and File Path
# ---------------------------------------------------------
target_directory = r"C:\Users\donca\Desktop\Desktop HP Envy x360 al 22Abr24\Docs Manuel\IBKR_Options"
file_name = "XSP_Options_Journal.xlsx"
full_file_path = os.path.join(target_directory, file_name)

if not os.path.exists(target_directory):
    os.makedirs(target_directory)
    print(f"Created new directory: {target_directory}")

# ---------------------------------------------------------
# Step 3: Define Columns and Create Base File
# ---------------------------------------------------------
columns =[
    "Tranche ID",                  # A
    "Open Date",                   # B
    "Ticker",                      # C
    "VIX at Entry",                # D
    "DTE at Entry",                # E
    "Short Strike",                # F
    "Long Strike",                 # G
    "Quantity",                    # H
    "Premium Collected (USD)",     # I
    "Collateral Locked (USD)",     # J (Formula)
    "Total Net Credit (USD)",      # K (Formula)
    "Target 50% Exit Price (USD)", # L (Formula)
    "Close Date",                  # M
    "Closing Price (USD)",         # N
    "Days in Trade",               # O (Formula)
    "Total P&L (USD)",             # P (Formula)
    "Return on Capital (ROC) %",   # Q (Formula)
    "Annualized ROC %",            # R (Formula)
    "Notes / Adjustments"          # S
]

try:
    # Generate the base file
    df = pd.DataFrame(columns=columns)
    df.to_excel(full_file_path, index=False, engine='openpyxl')
    
    # ---------------------------------------------------------
    # Step 4: Open the File with OpenPyXL to Inject Formulas
    # ---------------------------------------------------------
    wb = load_workbook(full_file_path)
    ws = wb.active
    ws.title = "Options Ledger"
    
    # Format Headers (Yellow Background, Bold, Centered)
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col_num in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        # Adjust column widths for readability
        ws.column_dimensions[get_column_letter(col_num)].width = 15

    # ---------------------------------------------------------
    # Step 5: Inject Dynamic Excel Formulas (Rows 2 to 100)
    # ---------------------------------------------------------
    # We use IF(ISBLANK) logic so the cells remain visually empty until data is entered.
    
    for i in range(2, 102):
        # J: Collateral Locked = (Short Strike - Long Strike) * 100 * Quantity
        ws[f'J{i}'] = f'=IF(OR(ISBLANK(F{i}),ISBLANK(G{i}),ISBLANK(H{i})),"",(F{i}-G{i})*100*H{i})'
        ws[f'J{i}'].number_format = '#,##0.00'

        # K: Total Net Credit = Premium Collected * 100 * Quantity
        ws[f'K{i}'] = f'=IF(OR(ISBLANK(I{i}),ISBLANK(H{i})),"",I{i}*100*H{i})'
        ws[f'K{i}'].number_format = '#,##0.00'

        # L: Target 50% Exit Price = Premium Collected / 2
        ws[f'L{i}'] = f'=IF(ISBLANK(I{i}),"",I{i}/2)'
        ws[f'L{i}'].number_format = '#,##0.00'

        # O: Days in Trade = Close Date - Open Date
        ws[f'O{i}'] = f'=IF(OR(ISBLANK(M{i}),ISBLANK(B{i})),"",M{i}-B{i})'
        ws[f'O{i}'].number_format = '0'

        # P: Total P&L = (Premium Collected - Closing Price) * 100 * Quantity
        ws[f'P{i}'] = f'=IF(OR(ISBLANK(N{i}),ISBLANK(I{i}),ISBLANK(H{i})),"",(I{i}-N{i})*100*H{i})'
        ws[f'P{i}'].number_format = '#,##0.00'

        # Q: ROC % = Total P&L / Collateral Locked
        ws[f'Q{i}'] = f'=IF(OR(ISBLANK(P{i}),ISBLANK(J{i}),J{i}=0),"",P{i}/J{i})'
        ws[f'Q{i}'].number_format = '0.00%'

        # R: Annualized ROC % = ROC % * (365 / Days in Trade)
        ws[f'R{i}'] = f'=IF(OR(ISBLANK(Q{i}),ISBLANK(O{i}),O{i}=0),"",Q{i}*(365/O{i}))'
        ws[f'R{i}'].number_format = '0.00%'

        # Format input currency columns (I and N)
        ws[f'I{i}'].number_format = '#,##0.00'
        ws[f'N{i}'].number_format = '#,##0.00'

    # Freeze the top header row so it stays visible when scrolling down
    ws.freeze_panes = "A2"
    
    # Save the updated workbook
    wb.save(full_file_path)
    
    print("\n========================================================")
    print(f"SUCCESS! The Automated Options Journal has been created.")
    print(f"Location: {full_file_path}")
    print("========================================================")

except PermissionError:
    print("\nERROR: Permission denied. Please close the Excel file if it is currently open and run the script again.")
except Exception as e:
    print(f"\nAn unexpected error occurred: {e}")